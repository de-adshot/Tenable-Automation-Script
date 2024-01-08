from openpyxl import load_workbook
import os
import glob
import pandas as pd

# Read a csv file to a dataframe


# Filter two columns


# Combine multiple dataframes
#df_combined = pd.concat([df1, df2, df3, ...])

# Output dataframe to excel file

def is_file_exsist(combined_file_name):
    check_file = os.path.isfile(combined_file_name)
    #print("Checking if file exsists: ", check_file)
    #print("file exsist checked")
    return check_file

def get_csv():
    path = input("Enter the location of the CSV(full directory): ")
    path = path + r'\*.csv'
    files = glob.glob(path)
    print('csv files are :',files)
    print("csv list obtained")
    return files




def newfile_to_extract_rows(combined_file_name,df,file):

    df.to_excel(combined_file_name, sheet_name="combined",index=False)
    print(file+" exported")
    

def exsisting(combined_file_name,df,file):
    wb_row = load_workbook(combined_file_name)
    sheet = wb_row["combined"]
    last_row = sheet.max_row
    with pd.ExcelWriter(combined_file_name, engine="openpyxl", mode="a",if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name="combined", startrow=last_row,index=False,header=False)
    print(file+" exported")
        

get_csv_list = get_csv()
combined_file_name = input("Enter the file name to export: ")
combined_file_name = combined_file_name + ".xlsx"
for file in get_csv_list:
    df = pd.read_csv(file)
    columns = ["Plugin Name","Severity", "IP Address","Protocol","Port","Synopsis","Description","Solution","See Also","Exploit?","Plugin Output","First Discovered"]
    df = df[columns]
    is_file_exsists = is_file_exsist(combined_file_name)
    
    if is_file_exsists == False:
        newfile_to_extract_rows(combined_file_name,df,file)
    
    else:
        exsisting(combined_file_name,df,file)