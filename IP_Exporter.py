import glob
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
 

def get_ip_list():
    path = '*.txt'
    files = glob.glob(path)
    #print('ip files are :',files)
    print("ip list obtained")
    return files



def is_file_exsist(ip_file_name):
    check_file = os.path.isfile(ip_file_name)
    #print("Checking if file exsists: ", check_file)
    print("file exsist checked")
    return check_file


def read_IP(ip_file):
    ip_list=[]
    with open(ip_file) as f: 
        for line in f:
            #print(line.strip())
            ip_list.append(line.strip())
    #print(ip_list)
    print("ip read")
    return(ip_list)

def is_df_empty_fun(filtered_df):
    #print(filtered_df)
    filtered_df_status = filtered_df.empty
    #print("Is DF empty? ", filtered_df_status)
    print("is df empty check")
    return filtered_df_status




def newfile_to_extract_rows(ip,data_file,exported_file,sheet_name):
    
    dataframe1 = pd.read_excel(data_file)
 
    #print(dataframe1)

    filtered_df = dataframe1[dataframe1["IP address"].str.fullmatch(ip)]
    is_df__empty = is_df_empty_fun(filtered_df)
    print("is df came out")
        
    if is_df__empty == False:
        print("if entered")
        filtered_df.to_excel(exported_file,sheet_name=sheet_name,index=False)
        
    else:
        print("else entered")
        missing_ip_write(ip,sheet_name)
        print("else came out")

    filtered_df.to_excel(exported_file,sheet_name=sheet_name,index=False)
    print("new file expract check")
    
    #print(type(filtered_df))

def missing_ip_write(ip,sheet_name):
    print("missing ip entered")
    file = "Ip_not_found.csv"
    wb_ip_not_found = is_file_exsist(file)
    header = 'Ip Address'
    if wb_ip_not_found == False:
        f = open(file, "a+")
        f.write(header+","+"Category"+"\n")
        f.write(str(ip)+","+sheet_name+"\n")
        f.close()

            
    else:
        f = open(file, "a")
        f.write(str(ip)+","+sheet_name+"\n")
        f.close()
    

    

def exsistingfile_to_extract_rows(ip,data_file,exported_file,sheet_name):
    wb = load_workbook(exported_file)
    
    if sheet_name in wb:
        #print(f"Sheet exist: ", sheet_name)
        ws = wb[sheet_name]
        dataframe1 = pd.read_excel(data_file)
        filtered_df = dataframe1[dataframe1["IP address"].str.fullmatch(ip)]
        is_df__empty = is_df_empty_fun(filtered_df)
        print("is df came out")
        
        if is_df__empty == False:
            print("if entered")
            filtered_df.to_excel(exported_file,sheet_name=sheet_name,index=False,header=False)
            for row in dataframe_to_rows(filtered_df, index=False,header=False):
                #print(row)
                ws.append(row)
            wb.save(exported_file)
        
        else:
            print("else entered")
            missing_ip_write(ip,sheet_name)
            print("else came out")
        
        
        
    else:
        wb.create_sheet(sheet_name)
        #print(f"Sheet exist: ", sheet_name)
        ws = wb[sheet_name]
        dataframe1 = pd.read_excel(data_file)
        filtered_df = dataframe1[dataframe1["IP address"].str.fullmatch(ip)]
        is_df__empty = is_df_empty_fun(filtered_df)
        print("is df came out")
        
        if is_df__empty == False:
            print("if entered")
            filtered_df.to_excel(exported_file,sheet_name=sheet_name,index=False,header=False)
            for row in dataframe_to_rows(filtered_df, index=False,header=False):
                #print(row)
                ws.append(row)
            wb.save(exported_file)
        
        else:
            print("else entered")
            missing_ip_write(ip,sheet_name)
            print("else came out")
                
            
            
        

data_file = input("Enter the filename of the xlsx master workbook(without extention): ")
exported_file = input("Enter the xlsx filename to write(without extention): ")

ip_file_list = get_ip_list()
exported_file = exported_file + '.xlsx'
data_file = data_file + '.xlsx'

for ip_file in ip_file_list:
    #print(ip_file)
    
    sheet_name = ip_file.strip('.txt')
    #print(ip_file_name)
    


    ip_list = read_IP(ip_file)
    
    for ip in ip_list:
        
        #print("Printing IP:",ip)
        is_file_exsists = is_file_exsist(exported_file)
    


        if is_file_exsists == False:
            newfile_to_extract_rows(ip,data_file,exported_file,sheet_name)
    
        else:
            exsistingfile_to_extract_rows(ip,data_file,exported_file,sheet_name)