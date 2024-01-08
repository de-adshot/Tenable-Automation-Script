print("***********************************************************************")
print("\n")
print("Nessus Report Generator V2.1 Dtd. 29 March 2021","\n")
print("Pls be ready with input file path and output file path", "\n")
print("To stop the program abruptly pls use Ctrl + d","\n")
print("***********************************************************************","\n")
#InputReportLoc = "C:\\Users\\HEMKUM\\Downloads\\data.xlsx"
#OutputReportLoc = 'C:\\Users\\HEMKUM\\Downloads\\nessusFinalReport.xlsx'

from pathlib import Path
import re
import openpyxl
from datetime import date, datetime
import time


for i in range (1,3):
    print("Enter the full path of the input file like in below example")
    print("Example : ")
    print("C:\\Users\\usr01\\Downloads\\data.xlsx")
    print("Enter : ")
    inputfile = input()
    print(inputfile,"\n")
    if inputfile:
        InputReportLoc = Path(inputfile)
        break
    else:
        print("Looks like you have not entered input file path", "\n")
        i = 2-i+1
        print("Pls enter the input file path, you have ", i, " more attempts","\n")

for i in range (1,3):
    print("Enter the full path of the output file similar to below example")
    print("Example : ")
    print("C:\\Users\\usr01\\Downloads\\clientNamexx_report.xlsx")
    print("Enter : ")
    outputfile = input()
    print(outputfile, "\n")
    if outputfile:
        OutputReportLoc = Path(outputfile)
        break
    else:
        print("Looks like you have not entered output file path","\n")
        i = 2 - i + 1
        print("Pls enter the output file path, you have ", i, " more attempts","\n")




start = time.time()

outputReport = openpyxl.Workbook()
sheet = outputReport.active
title = str(date.today())
sheet.title = title

inputReport = openpyxl.load_workbook(filename=InputReportLoc, read_only=True)
inputPage = inputReport.active
iRowCount = inputPage.max_row
iRowCount = iRowCount + 1
print("No. of rows read : ",inputPage.max_row)
print("No. of cols read : ",inputPage.max_column)

dlist = []


def readReport():

    for i in range(2, iRowCount):
        print("Printing I", i)
        for cell in inputPage[i]:
            dlist.append(cell.value)
            #print(cell.value)  #printing the whole cell value from plugin name
    return dlist


def createSheet():
    #create a excel sheet
    r1c1 = sheet.cell(row = 1, column = 1)
    r1c1.value = "IP"
    r1c2 = sheet.cell(row = 1, column = 2)
    r1c2.value = "Compliance_Number"
    r1c3 = sheet.cell(row = 1, column = 3)
    r1c3.value = "Compliance_Name"
    r1c4 = sheet.cell(row = 1, column = 4)
    r1c4.value = "Status"
    r1c5 = sheet.cell(row = 1, column = 5)
    r1c5.value = "Description"
    r1c6 = sheet.cell(row = 1, column = 6)
    r1c6.value = "Observed"
    r1c7 = sheet.cell(row = 1, column = 7)
    r1c7.value = "Expected"
    r1c8 = sheet.cell(row = 1, column = 8)
    r1c8.value = "Impact"
    r1c9 = sheet.cell(row = 1, column = 9)
    r1c9.value = "Solution"



def wrtRprtIPAddr(r):
    for l in range(1, dl + 1, 9):
        data = sheet.cell(row=r, column=1)
        data.value = dlist[l]
        r = r+1
        outputReport.save(OutputReportLoc)


def wrtRprtRest00(r):
    print('Result for row no: ',r,'\n')
    try:
        ComplianceNumber = re.search('(?P<c>^Check Name: +)(?P<cno>([A-Z]+_[A-Z]+\s[\d.]+)*)', dlist[l])
        print(ComplianceNumber,'\n')
        ComplianceNumber = (ComplianceNumber.group('cno'))
        data = sheet.cell(row=r, column=2)
        data.value = ComplianceNumber
        outputReport.save(OutputReportLoc)
    except Exception as msg:
        data = sheet.cell(row=r, column=2)
        data.value = "ComplianceNumber field syntax exceptional"
        outputReport.save(OutputReportLoc)
        print('row no: ', r)
        print(msg, "Check ComplianceNumber field",'\n')

def wrtRprtRest01(r):
    print('Result for row no: ',r,'\n')
    try:
        ComplianceName = re.search('(?P<c>:\s)(?P<cname>BAB_MBSS.*)', dlist[l])
        #BAB_MBSS\s(\d+.){3}\S+\s+(?P<c>\S+)\s(?P<cname>.*)
        print(ComplianceName, '\n')
        ComplianceName = (ComplianceName.group('cname'))
        data = sheet.cell(row=r, column=3)
        data.value = ComplianceName
        outputReport.save(OutputReportLoc)
    except Exception as msg:
        data = sheet.cell(row=r, column=3)
        data.value = 'ComplianceName field syntax exceptional'
        outputReport.save(OutputReportLoc)
        print('row no: ', r)
        print(msg, "Check ComplianceName field",'\n')

def wrtRprtRest02(r):
    print('Result for row no: ', r, '\n')
    try:
        Status = re.search("(?P<c>Result:\s)(?P<Status>\w+)(?P<d>\n)", dlist[l])
        print(Status,'\n')
        Status = (Status.group('Status'))
        data = sheet.cell(row=r, column=4)
        data.value = Status
        outputReport.save(OutputReportLoc)
    except Exception as msg:
        data = sheet.cell(row=r, column=4)
        data.value = "Status Field syntax exceptional"
        outputReport.save(OutputReportLoc)
        print('row no: ', r)
        print(msg, "Check Status field", '\n')

def wrtRprtRest03(r):
    print('Result for row no: ', r, '\n')
    try:
        xmatch = re.search("Information:\s", dlist[l])
        x = xmatch.span()[1]
        ymatch = re.search("\nImpact:", dlist[l])
        y = ymatch.span()[0]
        Desc = dlist[l][x:y]
        print('Desc', Desc,'\n')
        data = sheet.cell(row=r, column=5)
        data.value = Desc
        outputReport.save(OutputReportLoc)
    except Exception as msg:
        data = sheet.cell(row=r, column=5)
        data.value = "Desc. field syntax exceptional"
        outputReport.save(OutputReportLoc)
        print('row no: ', r)
        print(msg, "Check Desc field", '\n')

def wrtRprtRest04(r):
    print('Result for row no: ', r, '\n')
    if 'Actual Value' in dlist[l]:
        print('Actual Value True')
        try:
            #RmtVal = re.search('(?P<R>Actual Value:\n)(?P<RmtVal>(.*){0,8})', dlist[l])
            xmatch = re.search("Actual Value:", dlist[l])
            x = xmatch.span()[1]
            ymatch = re.search("\nPolicy Value:", dlist[l])
            y = ymatch.span()[0]
            RmtVal = dlist[l][x:y]
            print(RmtVal,'\n')
            #RmtVal = (RmtVal.group('RmtVal'))
            data = sheet.cell(row=r, column=6)
            data.value = RmtVal
            outputReport.save(OutputReportLoc)
        except Exception as msg:
            data = sheet.cell(row=r, column=6)
            data.value = "RmtVal field Syntax Exceptional"
            outputReport.save(OutputReportLoc)
            print('row no: ', r)
            print(msg, "Check RmtVal field", '\n')
    else:
        data = sheet.cell(row=r, column=6)
        data.value = "Actual Value field Not Available"
        outputReport.save(OutputReportLoc)
        print('row no: ', r)
        print("Actual Value field Not available", '\n')


def wrtRprtRest05(r):
    print('Result for row no: ', r)
    if 'Policy Value' in dlist[l]:
        print('Policy Value True')
        try:
            #PolVal = re.search('((?P<R>Policy Value:\n)(?P<PolVal>(.*\s){0,8})Actual)', dlist[l])
            xmatch = re.search("Policy Value:", dlist[l])
            x = xmatch.span()[1]
            ymatch = re.search("\nSolution:", dlist[l])
            y = ymatch.span()[0]
            PolVal = dlist[l][x:y]
            print(PolVal, '\n')
            #PolVal = (PolVal.group('PolVal'))
            data = sheet.cell(row=r, column=7)
            data.value = PolVal
            outputReport.save(OutputReportLoc)
        except Exception as msg:
            data = sheet.cell(row=r, column=7)
            data.value = "PolVal field syntax field exceptional"
            outputReport.save(OutputReportLoc)
            print('\n','row no: ', r)
            print(msg, "Check PolVal field", '\n')
    else:
        data = sheet.cell(row=r, column=7)
        data.value = "PolVal field Not Available"
        outputReport.save(OutputReportLoc)
        print('\n', 'row no: ', r)
        print("Policy Value Not available", '\n')


def wrtRprtRest06(r):
    print('\n')
    print('Result for row no: ', r)
    print(str(datetime.now()))
    if 'Impact' in dlist[l]:
        print('Impact True')
        try:
            amatch = re.search('(Impact:)', dlist[l])
            print(amatch)
            a = amatch.span()[1]
            bmatch = re.search('\nResult:', dlist[l])
            print(bmatch)
            b = bmatch.span()[0]
            Impact = dlist[l][a:b]
            print(Impact)
            data = sheet.cell(row=r, column=8)
            data.value = Impact
            outputReport.save(OutputReportLoc)
        except Exception as msg:
            data = sheet.cell(row=r, column=8)
            data.value = 'Impact field syntax exceptional'
            outputReport.save(OutputReportLoc)
            print('row no: ', r)
            print(msg, "Check Impact field", '\n')
    else:
        data = sheet.cell(row=r, column=8)
        data.value = 'Impact field Not Available'
        outputReport.save(OutputReportLoc)
        print('row no: ', r)
        print("Check Impact field", '\n')

def wrtRprtRest07(r):
    print('Result for row no: ', r)
    if 'Solution' in dlist[l]:
        print('Solution True')
        try:
            amatch = re.search('(Solution:\s)', dlist[l])
            print(amatch)
            a = amatch.span()[1]
            bmatch = re.search('(\nReference Information:)', dlist[l])
            print(bmatch)
            b = bmatch.span()[0]
            Impact = dlist[l][a:b]
            print(Impact)
            data = sheet.cell(row=r, column=9)
            data.value = Impact
            outputReport.save(OutputReportLoc)

            
        except Exception as msg:
            data = sheet.cell(row=r, column=9)
            data.value = "Solution field syntax exceptional"
            outputReport.save(OutputReportLoc)
            print('\n', 'row no: ', r)
            print(msg, "Check Solution field", '\n')
    else:
        data = sheet.cell(row=r, column=9)
        data.value = "Solution field not available"
        outputReport.save(OutputReportLoc)
        print('\n', 'row no: ', r)
        print("Solution field not available", '\n')



print("Reading input report ....")
print("Started at : ", str(datetime.now()))
readReport()
print("Ended at : ", str(datetime.now()), '\n')

dl = len(dlist)
print("Creating Finalreport(XL sheet) template ....")
print("Started at : ", str(datetime.now()))
createSheet()
print("Ended at : ", str(datetime.now()), '\n')

print("Extracting fields and writing into Finalreport(XL sheet) ....")
print("Started at : ", str(datetime.now()))
r=2
wrtRprtIPAddr(r)
for l in range(6, dl + 1, 9):
        wrtRprtRest00(r)
        wrtRprtRest01(r)
        wrtRprtRest02(r)
        wrtRprtRest03(r)
        wrtRprtRest04(r)
        wrtRprtRest05(r)
        wrtRprtRest06(r)
        wrtRprtRest07(r)
        r = r + 1
outputReport.save(OutputReportLoc)
print("Ended at : ", str(datetime.now()), '\n')

end = time.time()
totaltime = end - start
totaltimeSec = "{:.2f}".format(totaltime)
totaltimeMin = float(totaltimeSec)/60
totaltimeMin = "{:.2f}".format(totaltimeMin)
print("Report completed in : ", totaltimeSec, " seconds")
print("Report completed in : ", totaltimeMin, " minutes")

print('\n',"Please check the report at :", OutputReportLoc)


